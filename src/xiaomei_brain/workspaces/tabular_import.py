"""Import tabular files into a Workspace's durable business world."""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import re
from collections import Counter
from pathlib import Path
from typing import Any

from .business_service import BusinessWorldService
from .models import CollectionDefinition, DataSource, FieldDefinition
from .schema_resolver import SchemaAmbiguityError

MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 128


class TabularImportService:
    """Turn one CSV/TSV/XLSX snapshot into traceable Workspace records."""

    def __init__(self, business: BusinessWorldService) -> None:
        self.business = business

    def import_path(
        self,
        workspace_id: str,
        path: str | Path,
        *,
        source_name: str = "",
        sheet: str = "",
        collection_id: str = "",
        collection_name: str = "",
        collection_label: str = "",
        key_column: str = "",
        source_person_id: str = "",
        asset_id: str = "",
        session_id: str = "",
        turn_id: str = "",
    ) -> dict[str, Any]:
        source_path = Path(path).resolve()
        if not source_path.is_file():
            raise ValueError("The tabular source file is not available")
        suffix = source_path.suffix.lower()
        if suffix not in {".csv", ".tsv", ".xlsx"}:
            raise ValueError("Workspace import supports CSV, TSV and XLSX files")

        display_name = source_name.strip() or source_path.name
        selected_sheet, headers, rows = self._read(source_path, sheet)
        if not headers:
            raise ValueError("The tabular source has no columns")
        if not rows:
            raise ValueError("The tabular source has no data rows")
        digest = self._sha256(source_path)
        locator = f"attachment:{display_name.casefold()}"
        source = self.business.store.find_data_source(
            workspace_id, kind="file", locator=locator,
        )
        if source is None:
            source = self.business.create_data_source(
                workspace_id,
                kind="file",
                name=display_name,
                locator=locator,
                session_id=session_id,
                turn_id=turn_id,
            )

        external_ref = f"sha256:{digest}#sheet:{selected_sheet}"
        existing_observation = self.business.store.find_observation(
            source.id, external_ref,
        )
        if existing_observation is not None:
            return {
                "success": True,
                "duplicate": True,
                "message": "This exact file snapshot has already been imported",
                "data_source": self.business.data_source_snapshot(source),
                "observation": self.business.observation_snapshot_with_links(
                    existing_observation,
                ),
                "collection_id": existing_observation.resolved_collection_id,
                "row_count": len(rows),
                "created": 0,
                "updated": 0,
                "unchanged": len(rows),
            }

        collection = self._resolve_collection(
            workspace_id,
            source,
            headers,
            rows,
            collection_id=collection_id,
            collection_name=collection_name,
            collection_label=collection_label,
            session_id=session_id,
            turn_id=turn_id,
        )
        fields = self._ensure_fields(
            collection, headers, rows, session_id=session_id, turn_id=turn_id,
        )
        collection = self.business.require_collection(collection.id)
        column_fields = self._column_field_map(headers, fields)
        resolved_key = self._choose_key_column(headers, rows, key_column)
        prepared_rows = []
        for row_number, row in enumerate(rows, start=2):
            raw_values = {
                column_fields[header].name: self._coerce_value(
                    column_fields[header], row[index],
                )
                for index, header in enumerate(headers)
            }
            normalized = self.business._normalize_record_values(raw_values, fields)
            stable_key = self._stable_key(
                source,
                selected_sheet,
                row_number,
                key_field=column_fields.get(resolved_key) if resolved_key else None,
                normalized_values=normalized,
            )
            current = self.business.store.find_record_by_key(
                collection.id, stable_key,
            )
            key_field = column_fields.get(resolved_key) if resolved_key else None
            if current is None and key_field is not None:
                matches = self.business.store.query_records(
                    collection.id,
                    filters={key_field.id: normalized.get(key_field.id)},
                    limit=2,
                )
                if len(matches) > 1:
                    raise ValueError(
                        f"Multiple existing records match key column: {resolved_key}"
                    )
                current = matches[0] if matches else None
            is_unchanged = current is not None and all(
                current.values.get(field_id) == value
                for field_id, value in normalized.items()
            )
            prepared_rows.append((raw_values, current, stable_key, is_unchanged))

        observation = self.business.observe(
            workspace_id,
            data_source_id=source.id,
            source_person_id=source_person_id,
            external_ref=external_ref,
            content=(
                f"Imported {len(rows)} rows and {len(headers)} columns from "
                f"{display_name}{f' / {selected_sheet}' if selected_sheet else ''}"
            ),
            attributes={
                "format": suffix.lstrip("."),
                "sheet": selected_sheet,
                "headers": headers,
                "row_count": len(rows),
                "column_count": len(headers),
                "sha256": digest,
                "key_column": resolved_key,
            },
            asset_id=asset_id,
            session_id=session_id,
            turn_id=turn_id,
        )

        created = updated = unchanged = 0
        for raw_values, current, stable_key, is_unchanged in prepared_rows:
            if is_unchanged and current is not None:
                self.business.store.link_observation_to_record(
                    observation.id, collection.id, current.id,
                )
                unchanged += 1
                continue
            self.business.upsert_record(
                collection.id,
                values=raw_values,
                stable_key=stable_key,
                record_id=current.id if current is not None else "",
                expected_revision=current.revision if current is not None else None,
                business_intent=f"Import business data from {display_name}",
                person_id=source_person_id,
                session_id=session_id,
                turn_id=turn_id,
                observation_id=observation.id,
                notify=False,
            )
            if current is None:
                created += 1
            else:
                updated += 1

        resolved_observation = self.business.store.get_observation(observation.id)
        if resolved_observation is None:
            raise RuntimeError("Imported Observation disappeared")
        payload = {
            "workspace_id": workspace_id,
            "data_source": self.business.data_source_snapshot(source),
            "observation": self.business.observation_snapshot_with_links(
                resolved_observation,
            ),
            "collection": self.business.collection_snapshot(collection, fields),
            "source_name": display_name,
            "sheet": selected_sheet,
            "row_count": len(rows),
            "column_count": len(headers),
            "key_column": resolved_key,
            "schema_resolution": {
                "canonical_collection_id": collection.id,
                "fields_by_column": {
                    header: column_fields[header].id for header in headers
                },
            },
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "duplicate": False,
            "success": True,
        }
        self.business.publish_import_completed(
            workspace_id,
            payload,
            collection_id=collection.id,
            reason=f"Imported {display_name}",
            session_id=session_id,
            turn_id=turn_id,
        )
        return payload

    def _resolve_collection(
        self,
        workspace_id: str,
        source: DataSource,
        headers: list[str],
        rows: list[list[Any]],
        *,
        collection_id: str,
        collection_name: str,
        collection_label: str,
        session_id: str,
        turn_id: str,
    ) -> CollectionDefinition:
        if collection_id.strip():
            collection = self.business.require_collection(collection_id)
            if collection.workspace_id != workspace_id:
                raise ValueError("Collection does not belong to the Workspace")
            return collection
        previous = self.business.store.latest_resolved_observation(source.id)
        if previous is not None:
            reused = self.business.store.get_collection(previous.resolved_collection_id)
            if reused is not None and reused.workspace_id == workspace_id:
                return reused
        if not collection_name.strip() and not collection_label.strip():
            compatible = self._find_compatible_collection(workspace_id, headers)
            if compatible is not None:
                return compatible
        requested_identity = collection_name.strip() or collection_label.strip()
        if requested_identity:
            existing = self.business.schema.resolve_collection_identity(
                workspace_id,
                name=collection_name.strip(),
                label=collection_label.strip(),
            )
            if existing is not None:
                return existing
            machine_name = self._machine_name(requested_identity) or "imported_data"
        else:
            machine_name = self._available_collection_name(workspace_id, "")
        fields = self._infer_fields(headers, rows)
        collection, _ = self.business.create_collection(
            workspace_id,
            name=machine_name,
            label=collection_label.strip() or Path(source.name).stem,
            purpose=f"Business data imported from {source.name}",
            fields=fields,
            maturity="candidate",
            session_id=session_id,
            turn_id=turn_id,
        )
        return collection

    def _find_compatible_collection(
        self,
        workspace_id: str,
        headers: list[str],
    ) -> CollectionDefinition | None:
        return self.business.schema.resolve_import_collection(
            workspace_id,
            headers,
        )

    def _ensure_fields(
        self,
        collection: CollectionDefinition,
        headers: list[str],
        rows: list[list[Any]],
        *,
        session_id: str,
        turn_id: str,
    ) -> list[FieldDefinition]:
        existing = self.business.store.list_fields(collection.id)
        inferred = self._align_inferred_fields(
            self._infer_fields(headers, rows),
            existing,
        )
        _collection, fields = self.business.add_collection_fields(
            collection.id,
            fields=inferred,
            expected_revision=collection.revision,
            session_id=session_id,
            turn_id=turn_id,
        )
        return fields

    def _align_inferred_fields(
        self,
        inferred: list[dict[str, Any]],
        existing: list[FieldDefinition],
    ) -> list[dict[str, Any]]:
        """Let a durable schema interpret weak types inferred from one file."""
        aligned: list[dict[str, Any]] = []
        for proposal in inferred:
            resolved = None
            for identity in (proposal["label"], proposal["name"]):
                try:
                    resolved = self.business.schema.resolve_field(existing, identity)
                    break
                except SchemaAmbiguityError:
                    raise
                except ValueError:
                    continue
            if resolved is None:
                aligned.append(proposal)
                continue
            item = dict(proposal)
            item["data_type"] = resolved.data_type
            aligned.append(item)
        return aligned

    def _column_field_map(
        self,
        headers: list[str],
        fields: list[FieldDefinition],
    ) -> dict[str, FieldDefinition]:
        result: dict[str, FieldDefinition] = {}
        for header in headers:
            result[header] = self.business.schema.resolve_field(fields, header)
        return result

    def _available_collection_name(self, workspace_id: str, requested: str) -> str:
        base = self._machine_name(requested) or "imported_data"
        existing = {
            self.business.schema.identity_key(item.name)
            for item in self.business.store.list_collections(workspace_id)
        }
        if self.business.schema.identity_key(base) not in existing:
            return base
        for index in range(2, 10_000):
            candidate = f"{base}_{index}"
            if self.business.schema.identity_key(candidate) not in existing:
                return candidate
        raise ValueError("Unable to allocate a Collection name")

    @classmethod
    def _infer_fields(
        cls,
        headers: list[str],
        rows: list[list[Any]],
    ) -> list[dict[str, Any]]:
        names: Counter[str] = Counter()
        fields: list[dict[str, Any]] = []
        for index, header in enumerate(headers):
            base = cls._machine_name(header) or f"column_{index + 1}"
            names[base] += 1
            name = base if names[base] == 1 else f"{base}_{names[base]}"
            values = [row[index] for row in rows if index < len(row)]
            fields.append({
                "name": name,
                "label": header,
                "data_type": cls._infer_type(header, values),
                "required": False,
                "aliases": [],
            })
        return fields

    @staticmethod
    def _machine_name(value: str) -> str:
        text = re.sub(r"[^0-9A-Za-z_]+", "_", str(value).strip()).strip("_").lower()
        if text and text[0].isdigit():
            text = f"field_{text}"
        return text[:80]

    @classmethod
    def _infer_type(cls, header: str, values: list[Any]) -> str:
        present = [value for value in values if not cls._empty(value)]
        if not present:
            return "text"
        lowered = header.casefold()
        if all(cls._is_boolean(value) for value in present):
            return "boolean"
        if all(cls._is_number(value) for value in present):
            if any(token in lowered for token in ("金额", "价格", "收入", "成本", "amount", "price", "cost", "revenue")):
                return "money"
            if all(cls._is_integer(value) for value in present):
                return "integer"
            return "number"
        if all(cls._is_datetime(value) for value in present):
            return "datetime" if any(cls._has_time(value) for value in present) else "date"
        return "text"

    def _choose_key_column(
        self,
        headers: list[str],
        rows: list[list[Any]],
        requested: str,
    ) -> str:
        if requested.strip():
            requested_key = self.business.schema.identity_key(requested)
            match = next(
                (
                    header for header in headers
                    if self.business.schema.identity_key(header) == requested_key
                ),
                None,
            )
            if match is None:
                raise ValueError(f"Key column does not exist: {requested}")
            index = headers.index(match)
            values = [
                str(row[index]).strip()
                for row in rows if index < len(row) and not TabularImportService._empty(row[index])
            ]
            if len(values) != len(rows) or len(set(values)) != len(values):
                raise ValueError("Key column must contain a unique non-empty value in every row")
            return match
        candidates: list[tuple[int, int, str]] = []
        preferred = ("id", "编号", "编码", "单号", "邮箱", "email", "手机号", "phone", "名称", "name")
        for index, header in enumerate(headers):
            values = [str(row[index]).strip() for row in rows if index < len(row) and not TabularImportService._empty(row[index])]
            if len(values) != len(rows) or len(set(values)) != len(values):
                continue
            folded = header.casefold()
            rank = next((rank for rank, token in enumerate(preferred) if token in folded), len(preferred))
            candidates.append((rank, index, header))
        return min(candidates)[2] if candidates else ""

    @staticmethod
    def _stable_key(
        source: DataSource,
        sheet: str,
        row_number: int,
        *,
        key_field: FieldDefinition | None,
        normalized_values: dict[str, Any],
    ) -> str:
        if key_field is not None:
            value = normalized_values.get(key_field.id)
            return f"key:{key_field.id}:{str(value).strip()}"
        return f"source:{source.id}:sheet:{sheet}:row:{row_number}"

    @classmethod
    def _read(
        cls,
        path: Path,
        sheet_name: str,
    ) -> tuple[str, list[str], list[list[Any]]]:
        if path.suffix.lower() in {".csv", ".tsv"}:
            return "", *cls._read_csv(path)
        return cls._read_xlsx(path, sheet_name)

    @classmethod
    def _read_csv(cls, path: Path) -> tuple[list[str], list[list[Any]]]:
        text = None
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Unable to recognize the CSV encoding")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
        rows = csv.reader(text.splitlines(), dialect)
        try:
            headers = cls._headers(next(rows))
        except StopIteration as exc:
            raise ValueError("The tabular source is empty") from exc
        return headers, cls._limited_rows(rows, len(headers))

    @classmethod
    def _read_xlsx(
        cls,
        path: Path,
        sheet_name: str,
    ) -> tuple[str, list[str], list[list[Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX import requires openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            selected = sheet_name.strip() or workbook.sheetnames[0]
            if selected not in workbook.sheetnames:
                raise ValueError(f"Worksheet does not exist: {selected}")
            rows = workbook[selected].iter_rows(values_only=True)
            try:
                headers = cls._headers(list(next(rows)))
            except StopIteration as exc:
                raise ValueError("The worksheet is empty") from exc
            return selected, headers, cls._limited_rows(rows, len(headers))
        finally:
            workbook.close()

    @staticmethod
    def _headers(values: list[Any]) -> list[str]:
        if len(values) > MAX_IMPORT_COLUMNS:
            raise ValueError(f"Workspace import supports at most {MAX_IMPORT_COLUMNS} columns")
        used: Counter[str] = Counter()
        headers = []
        for index, value in enumerate(values, start=1):
            base = str(value or "").strip() or f"column_{index}"
            used[base] += 1
            headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
        return headers

    @staticmethod
    def _limited_rows(rows: Any, width: int) -> list[list[Any]]:
        result: list[list[Any]] = []
        for row in rows:
            if len(result) >= MAX_IMPORT_ROWS:
                raise ValueError(f"Workspace import supports at most {MAX_IMPORT_ROWS} rows at a time")
            values = list(row[:width])
            if not any(not TabularImportService._empty(value) for value in values):
                continue
            result.append(values + [None] * max(0, width - len(values)))
        return result

    @staticmethod
    def _empty(value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())

    @staticmethod
    def _is_boolean(value: Any) -> bool:
        if isinstance(value, bool):
            return True
        return isinstance(value, str) and value.strip().casefold() in {
            "true", "false", "yes", "no", "是", "否",
        }

    @staticmethod
    def _is_integer(value: Any) -> bool:
        if isinstance(value, bool):
            return False
        if isinstance(value, int):
            return True
        if isinstance(value, float):
            return value.is_integer()
        try:
            return float(str(value).strip().replace(",", "")).is_integer()
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _is_number(value: Any) -> bool:
        if isinstance(value, bool):
            return False
        try:
            float(str(value).strip().replace(",", ""))
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _is_datetime(value: Any) -> bool:
        if isinstance(value, (dt.date, dt.datetime)):
            return True
        if not isinstance(value, str):
            return False
        try:
            dt.datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
            return True
        except ValueError:
            return False

    @staticmethod
    def _has_time(value: Any) -> bool:
        if isinstance(value, dt.datetime):
            return True
        return isinstance(value, str) and ("T" in value or ":" in value)

    @classmethod
    def _coerce_value(cls, field: FieldDefinition, value: Any) -> Any:
        if cls._empty(value):
            return None
        kind = field.data_type
        if kind == "boolean":
            if isinstance(value, bool):
                return value
            return str(value).strip().casefold() in {"true", "yes", "是"}
        if kind == "integer":
            return int(float(str(value).strip().replace(",", "")))
        if kind in {"number", "money"}:
            return float(str(value).strip().replace(",", ""))
        if kind == "date":
            if isinstance(value, dt.datetime):
                return value.date().isoformat()
            if isinstance(value, dt.date):
                return value.isoformat()
            return dt.datetime.fromisoformat(
                str(value).strip().replace("Z", "+00:00"),
            ).date().isoformat()
        if kind == "datetime":
            if isinstance(value, dt.datetime):
                return value.isoformat()
            if isinstance(value, dt.date):
                return dt.datetime.combine(value, dt.time()).isoformat()
            return dt.datetime.fromisoformat(
                str(value).strip().replace("Z", "+00:00"),
            ).isoformat()
        if kind == "json":
            return value
        return str(value)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
