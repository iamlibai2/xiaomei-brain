"""Deterministic analysis for tabular attachments."""

from __future__ import annotations

import csv
import html
import math
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from xiaomei_brain.tools.base import Tool
from xiaomei_brain.tools.execution_context import current_tool_execution


MAX_ROWS = 50_000
MAX_COLUMNS = 200
MAX_GROUPS = 50


def _safe_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: Counter[str] = Counter()
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip() or f"column_{index}"
        used[base] += 1
        headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return headers[:MAX_COLUMNS]


def _read_csv(path: Path) -> tuple[list[str], list[list[Any]], bool]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 编码无法识别，请转换为 UTF-8 或 GB18030")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
    reader = csv.reader(text.splitlines(), dialect)
    try:
        headers = _safe_headers(next(reader))
    except StopIteration as exc:
        raise ValueError("数据文件为空") from exc
    rows: list[list[Any]] = []
    truncated = False
    for row in reader:
        if len(rows) >= MAX_ROWS:
            truncated = True
            break
        rows.append(list(row[:len(headers)]) + [""] * max(0, len(headers) - len(row)))
    return headers, rows, truncated


def _read_xlsx(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]], bool, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX 分析依赖 openpyxl 未安装") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = sheet_name.strip() or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {selected}")
        sheet = workbook[selected]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = _safe_headers(list(next(iterator)))
        except StopIteration as exc:
            raise ValueError("工作表为空") from exc
        rows: list[list[Any]] = []
        truncated = False
        for row in iterator:
            if len(rows) >= MAX_ROWS:
                truncated = True
                break
            values = list(row[:len(headers)])
            rows.append(values + [None] * max(0, len(headers) - len(values)))
        return headers, rows, truncated, selected
    finally:
        workbook.close()


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _profile(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, name in enumerate(headers):
        values = [row[index] if index < len(row) else None for row in rows]
        present = [value for value in values if value is not None and str(value).strip()]
        numbers = [number for value in present if (number := _number(value)) is not None]
        if not present:
            value_type = "empty"
        elif len(numbers) == len(present):
            value_type = "number"
        elif numbers:
            value_type = "mixed"
        else:
            value_type = "text"
        item: dict[str, Any] = {
            "name": name,
            "type": value_type,
            "non_empty": len(present),
            "missing": len(values) - len(present),
            "unique": len({str(value) for value in present}),
        }
        if numbers:
            item["numeric"] = {
                "count": len(numbers),
                "min": round(min(numbers), 6),
                "max": round(max(numbers), 6),
                "mean": round(statistics.fmean(numbers), 6),
                "median": round(statistics.median(numbers), 6),
            }
        result.append(item)
    return result


def _group_summary(
    headers: list[str],
    rows: list[list[Any]],
    group_by: str,
    value_columns: list[str],
) -> tuple[list[dict[str, Any]], bool]:
    if group_by not in headers:
        raise ValueError(f"分组列不存在: {group_by}")
    unknown = [name for name in value_columns if name not in headers]
    if unknown:
        raise ValueError(f"数值列不存在: {', '.join(unknown)}")
    group_index = headers.index(group_by)
    value_indexes = {name: headers.index(name) for name in value_columns}
    groups: dict[str, dict[str, list[float]]] = defaultdict(
        lambda: {name: [] for name in value_columns}
    )
    counts: Counter[str] = Counter()
    for row in rows:
        key = str(row[group_index] if group_index < len(row) else "").strip() or "(空值)"
        counts[key] += 1
        for name, index in value_indexes.items():
            number = _number(row[index] if index < len(row) else None)
            if number is not None:
                groups[key][name].append(number)
    ordered = sorted(counts, key=lambda key: (-counts[key], key))
    truncated = len(ordered) > MAX_GROUPS
    result = []
    for key in ordered[:MAX_GROUPS]:
        item: dict[str, Any] = {group_by: key, "count": counts[key]}
        for name, numbers in groups[key].items():
            item[name] = {
                "count": len(numbers),
                "sum": round(sum(numbers), 6),
                "mean": round(statistics.fmean(numbers), 6) if numbers else None,
            }
        result.append(item)
    return result, truncated


def _available_output_path(output_root: Path, name: str) -> Path:
    requested = output_root / name
    if not requested.exists():
        return requested
    for index in range(1, 10_000):
        candidate = output_root / f"{requested.stem} ({index}){requested.suffix}"
        if not candidate.exists():
            return candidate
    raise ValueError("无法分配图表输出文件名")


def _write_svg(
    groups: list[dict[str, Any]],
    group_by: str,
    value_column: str,
    chart_type: str,
    output_path: Path,
) -> None:
    points = []
    for item in groups[:20]:
        metric = item.get(value_column, {})
        value = metric.get("sum") if isinstance(metric, dict) else None
        if isinstance(value, (int, float)):
            points.append((str(item.get(group_by, "")), float(value)))
    if not points:
        raise ValueError("所选数值列没有可绘制的数据")
    width, height = 900, 520
    left, top, right, bottom = 80, 45, 30, 105
    plot_w, plot_h = width - left - right, height - top - bottom
    maximum = max(value for _, value in points) or 1.0
    minimum = min(0.0, min(value for _, value in points))
    span = maximum - minimum or 1.0
    zero_y = top + (maximum / span) * plot_h
    step = plot_w / max(1, len(points))
    elements = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        f'<text x="{left}" y="26" font-family="sans-serif" font-size="18" font-weight="600">{html.escape(value_column)} 按 {html.escape(group_by)} 汇总</text>',
        f'<line x1="{left}" y1="{zero_y:.1f}" x2="{width-right}" y2="{zero_y:.1f}" stroke="#c9ced6"/>',
    ]
    coords: list[tuple[float, float]] = []
    for index, (label, value) in enumerate(points):
        x = left + step * index + step / 2
        y = top + ((maximum - value) / span) * plot_h
        coords.append((x, y))
        if chart_type == "bar":
            bar_width = max(8, step * .58)
            bar_y = min(y, zero_y)
            bar_height = max(1, abs(zero_y - y))
            elements.append(f'<rect x="{x-bar_width/2:.1f}" y="{bar_y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" rx="3" fill="#4f7cff"/>')
        elements.append(f'<text x="{x:.1f}" y="{height-78}" transform="rotate(35 {x:.1f} {height-78})" font-family="sans-serif" font-size="11" fill="#59616e">{html.escape(label[:18])}</text>')
        elements.append(f'<text x="{x:.1f}" y="{max(16, y-8):.1f}" text-anchor="middle" font-family="sans-serif" font-size="10" fill="#3d4653">{value:g}</text>')
    if chart_type == "line":
        path = " ".join(("M" if index == 0 else "L") + f" {x:.1f} {y:.1f}" for index, (x, y) in enumerate(coords))
        elements.append(f'<path d="{path}" fill="none" stroke="#4f7cff" stroke-width="3"/>')
        elements.extend(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#4f7cff"/>' for x, y in coords)
    elements.append("</svg>")
    output_path.write_text("\n".join(elements), encoding="utf-8")


def analyze_data(
    attachment_id: str,
    sheet: str = "",
    group_by: str = "",
    value_columns: list[str] | None = None,
    chart_type: str = "none",
    output_name: str = "数据分析图表.svg",
) -> dict[str, Any]:
    context = current_tool_execution()
    if context is None:
        return {"error": "analyze_data 只能在 Agent 工具调用期间使用"}
    attachment = next(
        (item for item in context.attachments if str(item.get("id")) == attachment_id),
        None,
    )
    if attachment is None:
        return {"error": "当前执行现场中不存在这个附件"}
    source = Path(str(attachment.get("local_path") or ""))
    if not source.is_file():
        return {"error": "附件文件不可用"}
    try:
        suffix = source.suffix.lower()
        selected_sheet = ""
        if suffix in {".csv", ".tsv"}:
            headers, rows, truncated = _read_csv(source)
        elif suffix == ".xlsx":
            headers, rows, truncated, selected_sheet = _read_xlsx(source, sheet)
        else:
            return {"error": "目前只支持 CSV、TSV 和 XLSX 数据"}

        result: dict[str, Any] = {
            "success": True,
            "source_name": str(attachment.get("name") or source.name),
            "sheet": selected_sheet,
            "row_count": len(rows),
            "column_count": len(headers),
            "truncated": truncated,
            "columns": _profile(headers, rows),
        }
        values = list(dict.fromkeys(value_columns or []))
        if group_by:
            if not values:
                values = [item["name"] for item in result["columns"] if item["type"] in {"number", "mixed"}]
            summary, groups_truncated = _group_summary(headers, rows, group_by, values)
            result["group_summary"] = summary
            result["groups_truncated"] = groups_truncated
            if chart_type not in {"none", "bar", "line"}:
                raise ValueError("chart_type 只支持 none、bar 或 line")
            if chart_type != "none":
                if not values:
                    raise ValueError("生成图表需要至少一个数值列")
                safe_name = Path(output_name).name
                if safe_name != output_name or Path(safe_name).suffix.lower() != ".svg":
                    raise ValueError("output_name 必须是 .svg 普通文件名")
                output_root = Path(context.output_root or context.workspace_root).resolve()
                output_root.mkdir(parents=True, exist_ok=True)
                output_path = _available_output_path(output_root, safe_name).resolve()
                output_path.relative_to(output_root)
                _write_svg(summary, group_by, values[0], chart_type, output_path)
                result["chart"] = {
                    "type": chart_type,
                    "value_column": values[0],
                    "output_path": str(output_path),
                    "output_name": output_path.name,
                }
        return result
    except Exception as exc:
        return {"error": str(exc), "attachment_id": attachment_id}


def create_analyze_data_tool() -> Tool:
    return Tool(
        name="analyze_data",
        description=(
            "Profile a CSV/TSV/XLSX attachment from the current turn, calculate missing values "
            "and numeric statistics, optionally group rows and generate a bar or line SVG chart."
        ),
        parameters={
            "type": "object",
            "properties": {
                "attachment_id": {"type": "string"},
                "sheet": {"type": "string", "description": "Optional XLSX sheet name"},
                "group_by": {"type": "string", "description": "Optional grouping column"},
                "value_columns": {"type": "array", "items": {"type": "string"}},
                "chart_type": {"type": "string", "enum": ["none", "bar", "line"]},
                "output_name": {"type": "string", "description": "SVG file name when charting"},
            },
            "required": ["attachment_id"],
        },
        func=analyze_data,
        category="data",
    )
