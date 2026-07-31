"""Deterministic XLSX creation and non-destructive workbook updates."""

from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from xiaomei_brain.plugins.tools.document_spreadsheet.extractor import (
    SpreadsheetExtractor,
)


MAX_ROWS = 50_000
MAX_COLUMNS = 500
MAX_RANGE_CELLS = 100_000
INVALID_SHEET_NAME = re.compile(r"[\[\]:*?/\\]")


class SpreadsheetWriter:
    format_id = "spreadsheet"
    suffix = ".xlsx"
    writer_version = "1.0.0"

    def write(
        self,
        specification: dict[str, Any],
        output_path: Path,
        *,
        source_path: Path | None = None,
        asset_paths: dict[str, Path] | None = None,
    ) -> dict[str, Any]:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:
            raise ValueError("电子表格写入依赖 openpyxl 未安装") from exc

        workbook = None
        try:
            if source_path is not None:
                if source_path.suffix.lower() != self.suffix:
                    raise ValueError("Spreadsheet writer 只能修改 XLSX 附件")
                shutil.copy2(source_path, output_path)
                workbook = load_workbook(
                    output_path,
                    data_only=False,
                    keep_links=True,
                )
                operations = specification.get("operations")
                if not isinstance(operations, list) or not operations:
                    raise ValueError(
                        "修改电子表格时 specification.operations 不能为空"
                    )
                changed_cells = self._apply_operations(workbook, operations)
            else:
                sheets = specification.get("sheets")
                if not isinstance(sheets, list) or not sheets:
                    raise ValueError(
                        "创建电子表格时 specification.sheets 必须是非空数组"
                    )
                workbook = Workbook()
                workbook.remove(workbook.active)
                self._set_properties(workbook, specification.get("properties"))
                changed_cells = 0
                for sheet_spec in sheets:
                    changed_cells += self._add_sheet(workbook, sheet_spec)
            if not workbook.sheetnames:
                raise ValueError("电子表格至少需要一个工作表")
            workbook.save(output_path)
        finally:
            if workbook is not None:
                workbook.close()

        verified = load_workbook(
            output_path,
            read_only=False,
            data_only=False,
            keep_links=True,
        )
        try:
            formula_count, inspected_cells, truncated = self._formula_summary(
                verified
            )
            extraction = SpreadsheetExtractor().extract(output_path)
            preview = (
                extraction.sections[0].content[:1200]
                if extraction.sections
                else ""
            )
            return {
                "writer": self.format_id,
                "writer_version": self.writer_version,
                "validation": {
                    "valid": True,
                    "sheet_count": len(verified.sheetnames),
                    "sheet_names": list(verified.sheetnames),
                    "changed_cells": changed_cells,
                    "formula_count": formula_count,
                    "inspected_cells": inspected_cells,
                    "inspection_truncated": truncated,
                    "content_preview": preview,
                },
            }
        finally:
            verified.close()

    @staticmethod
    def _set_properties(workbook: Any, values: Any) -> None:
        if not isinstance(values, dict):
            return
        for key in ("title", "subject", "creator", "keywords", "description"):
            if key in values:
                setattr(workbook.properties, key, str(values[key]))

    @staticmethod
    def _sheet_name(value: Any) -> str:
        name = str(value or "").strip()
        if not name:
            raise ValueError("工作表名称不能为空")
        if len(name) > 31:
            raise ValueError("工作表名称不能超过 31 个字符")
        if INVALID_SHEET_NAME.search(name):
            raise ValueError("工作表名称包含 Excel 不允许的字符")
        return name

    @staticmethod
    def _sheet(workbook: Any, name: Any) -> Any:
        normalized = SpreadsheetWriter._sheet_name(name)
        if normalized not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {normalized}")
        return workbook[normalized]

    @staticmethod
    def _cell_bounds(reference: str) -> tuple[int, int, int, int]:
        from openpyxl.utils.cell import range_boundaries

        try:
            min_col, min_row, max_col, max_row = range_boundaries(reference)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"无效单元格或区域引用: {reference}") from exc
        if (
            min_row < 1
            or min_col < 1
            or max_row > MAX_ROWS
            or max_col > MAX_COLUMNS
        ):
            raise ValueError(
                f"单元格区域超出限制: 最大 {MAX_ROWS} 行、{MAX_COLUMNS} 列"
            )
        if (max_row - min_row + 1) * (max_col - min_col + 1) > MAX_RANGE_CELLS:
            raise ValueError(
                f"单次区域操作不能超过 {MAX_RANGE_CELLS} 个单元格"
            )
        return min_col, min_row, max_col, max_row

    @classmethod
    def _single_cell(cls, reference: str) -> tuple[int, int]:
        min_col, min_row, max_col, max_row = cls._cell_bounds(reference)
        if min_col != max_col or min_row != max_row:
            raise ValueError(f"需要单个单元格引用，收到: {reference}")
        return min_row, min_col

    @staticmethod
    def _typed_value(specification: Any) -> Any:
        if not isinstance(specification, dict):
            if isinstance(specification, (str, int, float, bool)) or specification is None:
                return specification
            raise ValueError("单元格值必须是字符串、数字、布尔值、null 或单元格对象")
        if "formula" in specification:
            formula = str(specification["formula"] or "").strip()
            if not formula:
                raise ValueError("单元格 formula 不能为空")
            return formula if formula.startswith("=") else "=" + formula
        value = specification.get("value")
        value_type = str(specification.get("type") or "").lower()
        if not value_type:
            return value
        try:
            if value_type == "date":
                return date.fromisoformat(str(value))
            if value_type == "datetime":
                return datetime.fromisoformat(str(value))
            if value_type == "time":
                return time.fromisoformat(str(value))
        except ValueError as exc:
            raise ValueError(f"无法按 {value_type} 解析单元格值: {value}") from exc
        raise ValueError(f"不支持的单元格值类型: {value_type}")

    @staticmethod
    def _hex_color(value: Any, field: str) -> str:
        color = str(value or "").strip().lstrip("#").upper()
        if len(color) == 6:
            color = "FF" + color
        if len(color) != 8 or any(ch not in "0123456789ABCDEF" for ch in color):
            raise ValueError(f"{field} 必须是 6 位或 8 位十六进制颜色")
        return color

    @classmethod
    def _apply_style(cls, cell: Any, values: Any) -> None:
        if not isinstance(values, dict):
            return
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        font_values = values.get("font")
        if isinstance(font_values, dict):
            current = copy(cell.font)
            color = (
                cls._hex_color(font_values["color"], "font.color")
                if "color" in font_values
                else current.color
            )
            cell.font = Font(
                name=font_values.get("name", current.name),
                size=font_values.get("size", current.sz),
                bold=font_values.get("bold", current.b),
                italic=font_values.get("italic", current.i),
                underline=font_values.get("underline", current.u),
                strike=font_values.get("strike", current.strike),
                color=color,
            )

        fill_value = values.get("fill")
        if fill_value is not None:
            color = cls._hex_color(fill_value, "fill")
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=color,
                bgColor=color,
            )

        alignment_values = values.get("alignment")
        if isinstance(alignment_values, dict):
            current = copy(cell.alignment)
            horizontal = alignment_values.get("horizontal", current.horizontal)
            vertical = alignment_values.get("vertical", current.vertical)
            if horizontal not in {
                None, "general", "left", "center", "right", "fill",
                "justify", "centerContinuous", "distributed",
            }:
                raise ValueError(f"不支持的水平对齐方式: {horizontal}")
            if vertical not in {
                None, "top", "center", "bottom", "justify", "distributed",
            }:
                raise ValueError(f"不支持的垂直对齐方式: {vertical}")
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=alignment_values.get(
                    "wrap_text",
                    current.wrap_text,
                ),
                text_rotation=alignment_values.get(
                    "text_rotation",
                    current.text_rotation,
                ),
            )

        border_values = values.get("border")
        if isinstance(border_values, dict):
            border_style = str(border_values.get("style") or "thin")
            color = cls._hex_color(
                border_values.get("color", "000000"),
                "border.color",
            )
            side = Side(style=border_style, color=color)
            cell.border = Border(left=side, right=side, top=side, bottom=side)

        if "number_format" in values:
            cell.number_format = str(values["number_format"])

    @classmethod
    def _set_cells(cls, sheet: Any, cells: Any) -> int:
        if not isinstance(cells, dict):
            raise ValueError("cells 必须是以单元格地址为键的对象")
        changed = 0
        for reference, cell_spec in cells.items():
            row, column = cls._single_cell(str(reference))
            cell = sheet.cell(row=row, column=column)
            cell.value = cls._typed_value(cell_spec)
            if isinstance(cell_spec, dict):
                cls._apply_style(cell, cell_spec.get("style"))
                if "number_format" in cell_spec:
                    cell.number_format = str(cell_spec["number_format"])
            changed += 1
        return changed

    @classmethod
    def _append_rows(cls, sheet: Any, rows: Any) -> int:
        if not isinstance(rows, list):
            raise ValueError("rows 必须是二维数组")
        is_empty = (
            sheet.max_row == 1
            and sheet.max_column == 1
            and sheet["A1"].value is None
        )
        existing_rows = 0 if is_empty else sheet.max_row
        if existing_rows + len(rows) > MAX_ROWS:
            raise ValueError(f"工作表不能超过 {MAX_ROWS} 行")
        changed = 0
        for offset, row_values in enumerate(rows, start=1):
            if not isinstance(row_values, list):
                raise ValueError("rows 中的每一行必须是数组")
            if len(row_values) > MAX_COLUMNS:
                raise ValueError(f"每行不能超过 {MAX_COLUMNS} 列")
            row_index = existing_rows + offset
            for column, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column)
                cell.value = cls._typed_value(value)
                if isinstance(value, dict):
                    cls._apply_style(cell, value.get("style"))
                    if "number_format" in value:
                        cell.number_format = str(value["number_format"])
                changed += 1
        return changed

    @classmethod
    def _style_range(cls, sheet: Any, reference: str, style: Any) -> int:
        min_col, min_row, max_col, max_row = cls._cell_bounds(reference)
        changed = 0
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                cls._apply_style(cell, style)
                changed += 1
        return changed

    @classmethod
    def _set_sheet_layout(cls, sheet: Any, values: Any) -> None:
        if not isinstance(values, dict):
            return
        from openpyxl.utils.cell import column_index_from_string

        freeze_panes = values.get("freeze_panes")
        if freeze_panes is not None:
            if freeze_panes:
                cls._single_cell(str(freeze_panes))
                sheet.freeze_panes = str(freeze_panes)
            else:
                sheet.freeze_panes = None

        auto_filter = values.get("auto_filter")
        if auto_filter is not None:
            if auto_filter:
                cls._cell_bounds(str(auto_filter))
                sheet.auto_filter.ref = str(auto_filter)
            else:
                sheet.auto_filter.ref = None

        widths = values.get("column_widths", {})
        if not isinstance(widths, dict):
            raise ValueError("column_widths 必须是对象")
        for column, width_value in widths.items():
            try:
                column_index = column_index_from_string(str(column).upper())
            except ValueError as exc:
                raise ValueError(f"无效列名: {column}") from exc
            if column_index > MAX_COLUMNS:
                raise ValueError(f"列超出限制: {column}")
            width = float(width_value)
            if not 0 < width <= 255:
                raise ValueError("列宽必须大于 0 且不超过 255")
            sheet.column_dimensions[str(column).upper()].width = width

        heights = values.get("row_heights", {})
        if not isinstance(heights, dict):
            raise ValueError("row_heights 必须是对象")
        for row, height_value in heights.items():
            try:
                row_index = int(row)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"无效行号: {row}") from exc
            if not 1 <= row_index <= MAX_ROWS:
                raise ValueError(f"行号超出限制: {row}")
            height = float(height_value)
            if not 0 < height <= 409:
                raise ValueError("行高必须大于 0 且不超过 409")
            sheet.row_dimensions[row_index].height = height

    @classmethod
    def _apply_sheet_spec(cls, sheet: Any, specification: Any) -> int:
        if not isinstance(specification, dict):
            raise ValueError("工作表 specification 必须是对象")
        changed = 0
        rows = specification.get("rows")
        if rows is not None:
            changed += cls._append_rows(sheet, rows)
        cells = specification.get("cells")
        if cells is not None:
            changed += cls._set_cells(sheet, cells)
        for reference in specification.get("merged_cells", []):
            cls._cell_bounds(str(reference))
            sheet.merge_cells(str(reference))
        styles = specification.get("styles", [])
        if not isinstance(styles, list):
            raise ValueError("styles 必须是数组")
        for style_spec in styles:
            if not isinstance(style_spec, dict):
                raise ValueError("styles 中的每一项必须是对象")
            changed += cls._style_range(
                sheet,
                str(style_spec.get("range") or ""),
                style_spec.get("style"),
            )
        cls._set_sheet_layout(sheet, specification)
        return changed

    @classmethod
    def _add_sheet(cls, workbook: Any, specification: Any) -> int:
        if not isinstance(specification, dict):
            raise ValueError("sheets 中的每一项必须是对象")
        name = cls._sheet_name(specification.get("name"))
        if name in workbook.sheetnames:
            raise ValueError(f"工作表已存在: {name}")
        sheet = workbook.create_sheet(name)
        return cls._apply_sheet_spec(sheet, specification)

    @classmethod
    def _apply_operations(cls, workbook: Any, operations: Iterable[Any]) -> int:
        changed = 0
        for operation in operations:
            if not isinstance(operation, dict):
                raise ValueError("Spreadsheet operation 必须是对象")
            kind = str(operation.get("type") or "")
            if kind == "set_cells":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                changed += cls._set_cells(sheet, operation.get("cells"))
            elif kind == "append_rows":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                changed += cls._append_rows(sheet, operation.get("rows"))
            elif kind == "style_range":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                changed += cls._style_range(
                    sheet,
                    str(operation.get("range") or ""),
                    operation.get("style"),
                )
            elif kind == "merge_cells":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                reference = str(operation.get("range") or "")
                cls._cell_bounds(reference)
                sheet.merge_cells(reference)
            elif kind == "set_sheet_layout":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                cls._set_sheet_layout(sheet, operation)
            elif kind == "add_sheet":
                changed += cls._add_sheet(workbook, operation.get("specification"))
            elif kind == "rename_sheet":
                sheet = cls._sheet(workbook, operation.get("sheet"))
                new_name = cls._sheet_name(operation.get("new_name"))
                if new_name in workbook.sheetnames and new_name != sheet.title:
                    raise ValueError(f"工作表已存在: {new_name}")
                sheet.title = new_name
            else:
                raise ValueError(f"不支持的 Spreadsheet operation: {kind}")
        return changed

    @staticmethod
    def _formula_summary(workbook: Any) -> tuple[int, int, bool]:
        formula_count = 0
        inspected = 0
        truncated = False
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if inspected >= MAX_RANGE_CELLS:
                        truncated = True
                        return formula_count, inspected, truncated
                    inspected += 1
                    if cell.data_type == "f":
                        formula_count += 1
        return formula_count, inspected, truncated
