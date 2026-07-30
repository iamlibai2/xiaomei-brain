"""Bounded XLSX extraction with one section per worksheet."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from xiaomei_brain.documents.models import DocumentExtraction, DocumentSection
from xiaomei_brain.documents.office_xml import bounded_text


MAX_ROWS = 10_000
MAX_COLUMNS = 200


def _value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ")


class SpreadsheetExtractor:
    extractor_id = "document_spreadsheet"
    extractor_version = "1.0.0"
    suffixes = (".xlsx",)
    mime_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    def extract(self, path: Path) -> DocumentExtraction:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(filename=path, read_only=True, data_only=False)
        except ImportError as exc:
            raise ValueError("表格解析依赖 openpyxl 未安装") from exc
        except Exception as exc:
            raise ValueError(f"无法解析电子表格: {path.name}") from exc
        try:
            sections = []
            for index, sheet in enumerate(workbook.worksheets, start=1):
                lines: list[str] = []
                truncated = False
                for row_index, row in enumerate(
                    sheet.iter_rows(max_row=MAX_ROWS, max_col=min(sheet.max_column, MAX_COLUMNS)),
                    start=1,
                ):
                    values = [_value(cell.value) for cell in row]
                    while values and not values[-1]:
                        values.pop()
                    if values:
                        lines.append("\t".join(values))
                    if row_index >= MAX_ROWS and sheet.max_row > MAX_ROWS:
                        truncated = True
                content = bounded_text("\n".join(lines) or "[工作表为空]")
                if truncated:
                    content += "\n[仅提取前 10000 行]"
                sections.append(DocumentSection(
                    key=f"sheet:{index}",
                    title=sheet.title,
                    content=content,
                    metadata={
                        "sheet": index,
                        "rows": sheet.max_row,
                        "columns": sheet.max_column,
                        "truncated": truncated,
                    },
                ))
        finally:
            workbook.close()
        return DocumentExtraction(
            extractor_id=self.extractor_id,
            extractor_version=self.extractor_version,
            sections=tuple(sections),
            metadata={"format": "xlsx", "sheet_count": len(sections)},
        )
