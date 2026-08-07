import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from xiaomei_brain.plugin.context import PluginContext
from xiaomei_brain.plugin.registry import PluginRegistry
from xiaomei_brain.plugins.tools.document_spreadsheet.adapter import (
    register as register_spreadsheet,
)
from xiaomei_brain.plugins.tools.document_io.tool import create_write_document_tool
from xiaomei_brain.tools.execution_context import bind_tool_execution


def _spreadsheet_registry() -> PluginRegistry:
    registry = PluginRegistry()
    context = PluginContext({}, "document_spreadsheet", "test", registry)
    register_spreadsheet(context)
    return registry


def test_spreadsheet_plugin_owns_writer_and_skill_directory():
    registry = _spreadsheet_registry()

    assert registry.get_document_writer("spreadsheet") is not None
    assert registry.list_document_writers() == ["spreadsheet"]
    skill_dirs = registry.get_skill_directories()
    assert len(skill_dirs) == 1
    assert (Path(skill_dirs[0]) / "SKILL.md").is_file()


def test_write_document_creates_styled_spreadsheet_with_formulas(tmp_path):
    registry = _spreadsheet_registry()
    tool = create_write_document_tool(registry)
    workspace = tmp_path / "workspace"
    outputs = workspace / "outputs"
    workspace.mkdir()
    spec = workspace / "sales.json"
    spec.write_text(json.dumps({
        "properties": {"creator": "Xiaomei", "title": "Sales"},
        "sheets": [{
            "name": "销售明细",
            "rows": [
                ["产品", "数量", "单价", "金额"],
                ["A", 2, 19.9, {
                    "formula": "B2*C2",
                    "number_format": "¥#,##0.00",
                }],
            ],
            "cells": {
                "A5": {"value": "统计日期"},
                "B5": {
                    "value": "2026-07-31",
                    "type": "date",
                    "number_format": "yyyy-mm-dd",
                },
            },
            "merged_cells": ["A7:D7"],
            "styles": [{
                "range": "A1:D1",
                "style": {
                    "font": {"bold": True, "color": "FFFFFF"},
                    "fill": "2F5597",
                    "alignment": {"horizontal": "center"},
                    "border": {"style": "thin", "color": "D9E2F3"},
                },
            }],
            "freeze_panes": "A2",
            "auto_filter": "A1:D2",
            "column_widths": {"A": 18, "D": 14},
            "row_heights": {"1": 24},
        }],
    }, ensure_ascii=False), encoding="utf-8")

    with bind_tool_execution(
        tool_call_id="call-sheet-create",
        tool_name="write_document",
        arguments={},
        artifact_callback=None,
        workspace_root=str(workspace),
        output_root=str(outputs),
    ):
        result = tool.execute(
            format="spreadsheet",
            specification_path="sales.json",
            output_name="sales.xlsx",
        )

    assert result["success"] is True
    assert result["validation"]["valid"] is True
    assert result["validation"]["sheet_names"] == ["销售明细"]
    assert result["validation"]["formula_count"] == 1
    workbook = load_workbook(outputs / "sales.xlsx", data_only=False)
    try:
        sheet = workbook["销售明细"]
        assert sheet["A1"].value == "产品"
        assert sheet["D2"].value == "=B2*C2"
        assert sheet["D2"].number_format == "¥#,##0.00"
        assert sheet["B5"].value.date() == date(2026, 7, 31)
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].fill.fgColor.rgb == "FF2F5597"
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == "A1:D2"
        assert sheet.column_dimensions["A"].width == 18
        assert sheet.row_dimensions[1].height == 24
        assert "A7:D7" in {str(item) for item in sheet.merged_cells.ranges}
        assert workbook.properties.creator == "Xiaomei"
    finally:
        workbook.close()


def test_write_document_modifies_copy_and_keeps_workbook_structure(tmp_path):
    registry = _spreadsheet_registry()
    tool = create_write_document_tool(registry)
    workspace = tmp_path / "workspace"
    outputs = workspace / "outputs"
    workspace.mkdir()
    source = tmp_path / "source.xlsx"
    original = Workbook()
    data = original.active
    data.title = "数据"
    data.append(["产品", "数量", "单价", "金额"])
    data.append(["A", 2, 10, "=B2*C2"])
    original.save(source)
    original.close()
    original_bytes = source.read_bytes()

    spec = workspace / "update.json"
    spec.write_text(json.dumps({
        "operations": [
            {
                "type": "set_cells",
                "sheet": "数据",
                "cells": {
                    "B2": 3,
                    "D2": {"formula": "B2*C2", "number_format": "0.00"},
                },
            },
            {
                "type": "append_rows",
                "sheet": "数据",
                "rows": [["B", 5, 9.9, {"formula": "B3*C3"}]],
            },
            {
                "type": "style_range",
                "sheet": "数据",
                "range": "A1:D1",
                "style": {"font": {"bold": True}, "fill": "D9EAF7"},
            },
            {
                "type": "set_sheet_layout",
                "sheet": "数据",
                "freeze_panes": "A2",
                "auto_filter": "A1:D3",
                "column_widths": {"A": 20},
            },
            {
                "type": "add_sheet",
                "specification": {
                    "name": "说明",
                    "rows": [["字段", "含义"], ["金额", "数量乘以单价"]],
                },
            },
            {
                "type": "rename_sheet",
                "sheet": "说明",
                "new_name": "数据说明",
            },
        ],
    }, ensure_ascii=False), encoding="utf-8")
    attachment = {
        "id": "sheet-source",
        "name": "source.xlsx",
        "kind": "document",
        "local_path": str(source),
    }

    with bind_tool_execution(
        tool_call_id="call-sheet-update",
        tool_name="write_document",
        arguments={},
        artifact_callback=None,
        attachments=(attachment,),
        workspace_root=str(workspace),
        output_root=str(outputs),
    ):
        result = tool.execute(
            format="spreadsheet",
            specification_path="update.json",
            output_name="updated.xlsx",
            source_attachment_id="sheet-source",
        )

    assert result["success"] is True
    assert result["validation"]["formula_count"] == 2
    assert source.read_bytes() == original_bytes
    updated = load_workbook(outputs / "updated.xlsx", data_only=False)
    try:
        assert updated.sheetnames == ["数据", "数据说明"]
        sheet = updated["数据"]
        assert sheet["B2"].value == 3
        assert sheet["D2"].value == "=B2*C2"
        assert sheet["A3"].value == "B"
        assert sheet["D3"].value == "=B3*C3"
        assert sheet["A1"].font.bold is True
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == "A1:D3"
        assert updated["数据说明"]["B2"].value == "数量乘以单价"
    finally:
        updated.close()


def test_spreadsheet_writer_rejects_invalid_sheet_and_oversized_range(tmp_path):
    registry = _spreadsheet_registry()
    tool = create_write_document_tool(registry)
    workspace = tmp_path / "workspace"
    outputs = workspace / "outputs"
    workspace.mkdir()
    invalid_name = workspace / "invalid-name.json"
    invalid_name.write_text(json.dumps({
        "sheets": [{"name": "财务/数据", "rows": [["A"]]}],
    }, ensure_ascii=False), encoding="utf-8")
    invalid_range = workspace / "invalid-range.json"
    invalid_range.write_text(json.dumps({
        "sheets": [{
            "name": "数据",
            "rows": [["A"]],
            "styles": [{
                "range": "A1:XFD1048576",
                "style": {"font": {"bold": True}},
            }],
        }],
    }, ensure_ascii=False), encoding="utf-8")

    with bind_tool_execution(
        tool_call_id="call-sheet-invalid",
        tool_name="write_document",
        arguments={},
        artifact_callback=None,
        workspace_root=str(workspace),
        output_root=str(outputs),
    ):
        name_result = tool.execute(
            format="spreadsheet",
            specification_path="invalid-name.json",
            output_name="invalid-name.xlsx",
        )
        range_result = tool.execute(
            format="spreadsheet",
            specification_path="invalid-range.json",
            output_name="invalid-range.xlsx",
        )

    assert "error" in name_result
    assert "error" in range_result
    assert not list(outputs.glob("*.xlsx"))
